from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
FLOW_OUT = ROOT / "docs" / "Documind_v0.4_flows_business_rules.xlsx"
API_OUT = ROOT / "docs" / "Documind_API_Contract_v0.4.xlsx"

FLOWS = {
    "MF01": {
        "name": "Xác thực, hồ sơ, email và thanh toán tài nguyên",
        "steps": [
            ("MF01-BR01", "Đăng ký tài khoản", "User", "POST /api/auth/register", "Xác thực Firebase bearer token; tạo/cập nhật user local ở trạng thái chờ xác minh email; mật khẩu không lưu trong DB ứng dụng.", "auth.controller.ts -> AuthService.register()", "RegisterView.tsx / mobile register", "Đã có"),
            ("MF01-BR02", "Tạo link xác thực email", "System", "Firebase Admin generateEmailVerificationLink()", "Link một lần, dùng action URL của frontend; user EMAIL_PASSWORD chưa verify không được đăng nhập.", "mail/auth-email.service.ts", "Không gọi trực tiếp", "Đã có"),
            ("MF01-BR03", "Gửi mail xác thực qua SMTP", "Mail system", "SMTP transport nội bộ, không mở public API", "Dùng SMTP_HOST/PORT/SECURE/USER/PASSWORD; FROM thuộc domain đã xác minh; timeout/retry hữu hạn; không log secret hoặc action code.", "mail/mail.service.ts (cần đổi transport)", "Mail client của user", "Cần bổ sung SMTP"),
            ("MF01-BR04", "Xử lý link xác thực", "User", "Firebase action code tại /verify-email", "Action code dùng một lần và có hạn; sau verify, lần login tiếp theo đồng bộ emailVerified và kích hoạt user.", "Firebase + AuthService.firebaseLogin()", "VerifyEmailView.tsx", "Đã có"),
            ("MF01-BR05", "Đăng nhập và đồng bộ", "User", "POST /api/auth/firebase-login", "Firebase xác minh identity; PostgreSQL quyết định role/status; BLOCKED/INACTIVE hoặc email chưa verify bị từ chối.", "auth.controller.ts -> AuthService.firebaseLogin()", "LoginView.tsx / mobile login", "Đã có"),
            ("MF01-BR06", "Hydrate phiên hiện tại", "User", "GET /api/auth/me", "Chỉ trả profile/role/status hiện tại; role không lấy từ client/custom claim.", "auth.controller.ts -> AuthService.getCurrentUser()", "AuthProvider/Profile", "Đã có"),
            ("MF01-BR07", "Quên mật khẩu", "Guest", "POST /api/auth/forgot-password", "Luôn phản hồi 204 để chống dò email; Firebase Admin tạo reset link nếu tài khoản tồn tại.", "auth.controller.ts -> AuthEmailService", "ForgotPasswordView/mobile", "Đã có"),
            ("MF01-BR08", "Gửi mail reset qua SMTP", "Mail system", "SMTP transport nội bộ", "Template reset riêng; link một lần/có hạn; không tiết lộ email có tồn tại; lỗi SMTP trả 503 nội bộ nhưng thông điệp client vẫn an toàn.", "mail/mail.service.ts (cần đổi transport)", "Mail client của user", "Cần bổ sung SMTP"),
            ("MF01-BR09", "Đăng xuất", "User", "POST /api/auth/logout", "Xóa secure session cookie; frontend/mobile đồng thời sign out Firebase.", "auth.controller.ts", "AuthProvider/layout", "Đã có"),
            ("MF01-BR10", "Xem/cập nhật hồ sơ", "User", "GET/PATCH /api/users/profile; POST /api/users/avatar", "Chỉ user hiện tại; whitelist field; email do Firebase quản lý; avatar hợp lệ và không lộ R2 key.", "users.controller.ts -> UsersService", "ProfileView/mobile profile", "Đã có"),
            ("MF01-BR11", "Xem bundle và ví", "Guest/User", "GET /api/subscription/plans; GET /api/subscription/current", "Bundle tài nguyên cộng dồn; ví tính expiry/storage/upload/AI; hết hạn về Free baseline.", "payments.controller.ts -> PaymentsService", "SubscriptionView", "Đã có"),
            ("MF01-BR12", "Tạo checkout", "User", "POST /api/payments/checkout", "Đóng băng giá/quyền lợi trong order; ký HMAC phía server; redirect không tự cấp tài nguyên.", "payments.controller.ts -> PaymentsService", "SubscriptionView", "Đã có"),
            ("MF01-BR13", "Nhận IPN và cộng quyền", "SePay", "POST /api/payments/sepay/ipn", "Verify API key/invoice/amount; PAID và EntitlementTransaction trong cùng transaction; idempotent đúng một lần.", "payments.controller.ts -> PaymentsService", "SePay webhook", "Đã có"),
            ("MF01-BR14", "Theo dõi/lịch sử/hủy", "User", "GET /api/payments/:invoiceNumber; GET /api/payments/history; POST /api/payments/:invoiceNumber/status", "Owner only; PENDING có thể reconcile; client không đổi PAID; reversal không xóa quyền lợi từ order khác.", "payments.controller.ts -> PaymentsService", "SubscriptionView", "Đã có"),
        ],
    },
    "MF02": {
        "name": "Tải tài liệu và xử lý AI nền",
        "steps": [
            ("MF02-BR01", "FE gửi file và metadata", "User", "POST /api/documents", "Multipart; PDF/DOCX/PPTX/XLSX; mặc định <=10 MB; subject/category thuộc user; kiểm storage/upload quota.", "documents.controller.ts -> DocumentsService.upload()", "UploadDocumentView/mobile", "Đã có"),
            ("MF02-BR02", "Upload Cloudflare R2", "System", "Nội bộ sau POST /documents", "Object private, key theo users/{ownerId}/documents/{documentId}; không lộ credential/key; DB lỗi thì best-effort cleanup.", "storage.service.ts -> uploadObject()", "Không gọi trực tiếp", "Đã có"),
            ("MF02-BR03", "Tạo metadata/content", "System", "Nội bộ sau upload", "Document và DocumentContent tạo transactionally; extractionStatus=PENDING/progress=0; PUBLIC có moderation=PENDING.", "documents.service.ts + Prisma", "Không gọi trực tiếp", "Đã có"),
            ("MF02-BR04", "Xếp job extraction", "System/User", "Tự động; POST /api/documents/:id/extract để retry", "Không chạy hai job đồng thời; tạo jobId; retry/concurrency hữu hạn; worker đọc file từ R2.", "ContentExtractionService.startExtraction()", "Library retry", "Đã có"),
            ("MF02-BR05", "Extract/OCR/chunk/embed", "Worker", "Pipeline nội bộ", "Chọn extractor theo loại; OCR fallback; lưu text/quality/provenance; chunk có traceability; lỗi -> FAILED an toàn.", "content-extraction.service.ts", "Không gọi trực tiếp", "Đã có"),
            ("MF02-BR06", "Đọc content", "User", "GET /api/documents/:id/content", "Phải có quyền và COMPLETED; chưa sẵn sàng trả 409.", "document-content.controller.ts", "Document/chat UI", "Đã có"),
            ("MF02-BR07", "Theo dõi status", "User", "GET /api/documents/:id/extraction-status", "Trả PENDING/PROCESSING/COMPLETED/FAILED, progress, jobId và lỗi an toàn.", "document-content.controller.ts", "Library polling", "Đã có"),
            ("MF02-BR08", "Quét moderation risk", "System", "Pipeline nội bộ", "Scanner chỉ flag/priority/keyword/context; tuyệt đối không tự APPROVED tài liệu PUBLIC.", "moderation-scanner.service.ts", "Admin queue", "Đã có"),
            ("MF02-BR09", "Preview/download", "User", "GET /api/documents/:id/preview; GET /api/documents/:id/download", "Kiểm quyền trước signed URL; expiry ngắn; download ghi counter/log; Office có conversion fallback.", "documents.controller.ts -> DocumentsService", "LibraryView", "Đã có"),
        ],
    },
    "MF03": {
        "name": "Hỏi đáp AI với RAG",
        "steps": [
            ("MF03-BR01", "Hỏi trên một tài liệu", "User", "POST /api/chat/ask-document", "Document accessible và extraction COMPLETED; kiểm/tiêu AI credit atomically; trả answer + sources + suggestions.", "ai-chatbot.controller.ts -> AiChatbotService", "AskDocumentView", "Đã có"),
            ("MF03-BR02", "Hỏi trên thư viện", "User", "POST /api/chat/ask-library", "Scope chỉ owned + saved còn truy cập; content chưa ready có thể 409; không lộ private document người khác.", "ai-chatbot.controller.ts -> AiChatbotService", "AskLibrary/AiChatbotView", "Đã có"),
            ("MF03-BR03", "Streaming SSE", "User", "POST /api/chat/ask-library/stream; POST /api/chat/ask-document/stream", "Event retrieving -> generating -> sources -> delta -> verifying -> done/error; cùng auth/quota/persistence với non-stream.", "ai-chatbot.controller.ts", "Chat UI", "Đã có"),
            ("MF03-BR04", "Retrieval/chọn context", "System", "Pipeline nội bộ", "Xếp hạng chunk trong scope; citation phải trace documentId/chunk/snippet; không tạo nguồn ngoài context.", "ai-chatbot.service.ts + services", "Không gọi trực tiếp", "Đã có"),
            ("MF03-BR05", "Lưu session/message", "System", "Nội bộ trong ask", "Session thuộc user; lưu user/assistant message và delivery status; failed không được biểu diễn như complete.", "AiChatbotService", "Chat history", "Đã có"),
            ("MF03-BR06", "Mở lịch sử chat", "User", "GET /api/chat/sessions; GET /api/chat/sessions/:id; GET /api/chat/messages/:sessionId", "Owner session only; phân trang và thứ tự ổn định.", "ai-chatbot.controller.ts", "AiChatbotView", "Đã có"),
        ],
    },
    "MF04": {
        "name": "Chia sẻ cộng đồng và lưu vào thư viện",
        "steps": [
            ("MF04-BR01", "Chuyển sang PUBLIC", "Owner", "PUT /api/documents/:id/visibility", "Owner only; PUBLIC bắt buộc moderation=PENDING; PRIVATE không vào moderation/community.", "documents.controller.ts -> DocumentsService", "LibraryView", "Đã có"),
            ("MF04-BR02", "Xem cộng đồng", "Guest/User", "GET /api/community/documents", "Backend chỉ query PUBLIC + APPROVED + ACTIVE; optional auth chỉ bổ sung saved flag.", "community.controller.ts -> CommunityService", "CommunityView/mobile", "Đã có"),
            ("MF04-BR03", "Xem chi tiết", "Guest/User", "GET /api/community/documents/:id", "Cùng publication predicate; pending/rejected/hidden trả 404.", "community.controller.ts -> CommunityService", "CommunityView", "Đã có"),
            ("MF04-BR04", "Preview cộng đồng", "User", "GET /api/community/documents/:id/preview", "Yêu cầu auth và tài liệu vẫn public-approved-active; signed URL ngắn hạn.", "community.controller.ts -> CommunityService", "CommunityView", "Đã có"),
            ("MF04-BR05", "Lưu vào thư viện", "User", "POST /api/community/documents/:id/save", "Idempotent; unique user-document; chỉ lưu tài liệu accessible; được đưa vào Ask Library.", "community.controller.ts -> CommunityService", "Community/SavedView", "Đã có"),
            ("MF04-BR06", "Bỏ lưu", "User", "DELETE /api/community/documents/:id/save", "Idempotent; chỉ xóa relation của user; trả 204.", "community.controller.ts -> CommunityService", "Community/SavedView", "Đã có"),
            ("MF04-BR07", "Danh sách đã lưu", "User", "GET /api/saved-documents", "Chỉ saved relation user và document còn quyền; phân trang deterministic.", "saved-documents.controller.ts", "SavedView", "Đã có"),
        ],
    },
    "MF05": {
        "name": "Quản trị người dùng và kiểm duyệt tài liệu",
        "steps": [
            ("MF05-BR01", "User gửi duyệt", "Owner", "PUT /api/documents/:id/visibility", "PUBLIC -> PENDING; scanner không tự duyệt; owner có thể về PRIVATE.", "documents.controller.ts", "LibraryView", "Đã có"),
            ("MF05-BR02", "Admin xem hàng chờ", "Admin", "GET /api/admin/documents", "FirebaseAuthGuard + RolesGuard(ADMIN); chỉ PUBLIC; filter status/flag/keyword; ưu tiên risk/submittedAt.", "admin-documents.controller.ts", "AdminDocumentsView", "Đã có"),
            ("MF05-BR03", "Admin preview", "Admin", "GET /api/admin/documents/:id/preview", "Chỉ PUBLIC; preview không đổi trạng thái; signed URL ngắn hạn.", "admin-documents.controller.ts", "AdminDocumentsView", "Đã có"),
            ("MF05-BR04", "Admin duyệt", "Admin", "PUT /api/admin/documents/:id/approve", "Set APPROVED + ACTIVE; clear reason; reviewer/time; audit + notification owner.", "admin-documents.controller.ts", "AdminDocumentsView", "Đã có"),
            ("MF05-BR05", "Admin từ chối", "Admin", "PUT /api/admin/documents/:id/reject", "Reason trim bắt buộc; REJECTED + HIDDEN; reviewer/time; audit + notification owner.", "admin-documents.controller.ts", "AdminDocumentsView", "Đã có"),
            ("MF05-BR06", "Quản trị user", "Admin", "GET /api/admin/users; PATCH /api/admin/users/:id/status", "ADMIN role từ DB; không lộ credential; status transition có audit; BLOCKED/INACTIVE bị chặn protected API.", "admin-users.controller.ts", "AdminUsersView", "Đã có"),
            ("MF05-BR07", "Audit/notification/community gate", "System/Admin/User", "GET /api/admin/logs/audit; GET/PATCH /api/notifications; GET /api/community/documents", "Audit actor/action/resource; notification owner-only; community gate PUBLIC+APPROVED+ACTIVE là lớp bảo vệ cuối.", "audit-log/notifications/community modules", "AdminDashboard/AppLayout", "Đã có"),
        ],
    },
}

MEMBERS = {
    "MF01": ("Thống", "Flutter core/auth/profile/subscription; Firebase iOS/Android; SMTP email integration", "mobile/lib/features/auth; profile; subscription; mobile/lib/core/api_client.dart"),
    "MF02": ("Đăng Khôi", "Flutter upload, document library, extraction polling/retry, preview/download", "mobile/lib/features/documents/documents_screen.dart"),
    "MF03": ("Huân Minh", "Flutter AI chat/RAG, ask-document, ask-library, SSE/history", "mobile/lib/features/chat/chat_screen.dart"),
    "MF04": ("Phú Vinh", "Flutter community, public detail/preview, save/unsave và saved library", "mobile/lib/features/community/community_hub_screen.dart; documents_screen.dart"),
    "MF05": ("Đức Nguyên", "Flutter admin users, moderation queue, approve/reject, dashboard/log/report", "mobile/lib/features/admin/admin_screen.dart"),
}

FLOW_HEADERS = ["Rule ID", "Bước trong main flow", "Actor", "API / Trigger", "Business rule", "Controller / Logic", "Web FE / Consumer", "Trạng thái", "Thành viên", "Mobile App theo flow"]
API_HEADERS = ["API ID", "Main Flow", "Method", "Endpoint", "Auth / Actor", "Request", "Success", "Status", "Errors / Validation", "Business rule", "Controller / Service", "Web consumer", "Thành viên", "Mobile mapping", "Mobile status"]

API_ROWS = [
    ("MF01", "POST", "/api/auth/register", "Firebase bearer", "Register DTO + bearer", "User pending verification", "200", "400/401/409/503", "Tạo local user; gọi mail verification; không lưu password", "auth.controller.ts -> AuthService", "RegisterView/mobile"),
    ("MF01", "POST", "/api/auth/firebase-login", "Firebase bearer", "Bearer", "User + session cookie", "200", "401/403", "Email verified; DB ACTIVE; sync Firebase/local", "auth.controller.ts -> AuthService", "LoginView/mobile"),
    ("MF01", "GET", "/api/auth/me", "User", "—", "Current user", "200", "401/403", "Role/status từ PostgreSQL", "auth.controller.ts -> AuthService", "AuthProvider"),
    ("MF01", "POST", "/api/auth/forgot-password", "Public", "email", "No content", "204", "400/503", "Anti-enumeration; generate Firebase reset link; dispatch email", "auth.controller.ts -> AuthEmailService", "ForgotPasswordView/mobile"),
    ("MF01", "POST", "/api/auth/logout", "Session", "—", "No content", "204", "—", "Clear secure cookie; client signOut Firebase", "auth.controller.ts", "AuthProvider"),
    ("MF01", "GET/PATCH", "/api/users/profile", "User", "UpdateProfileDto for PATCH", "Profile", "200", "400/401/403", "Current user only; whitelist; email read-only", "users.controller.ts -> UsersService", "ProfileView"),
    ("MF01", "POST", "/api/users/avatar", "User", "multipart image", "Avatar metadata", "201", "400/413/415", "Validate image; replace safely; hide storage key", "users.controller.ts -> UsersService", "ProfileView"),
    ("MF01", "GET", "/api/subscription/plans", "Public", "—", "Bundle[]", "200", "—", "Additive resource bundles", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "GET", "/api/subscription/current", "User", "—", "Entitlement wallet", "200", "401/403", "Compute remaining resources and expiry", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "POST", "/api/payments/checkout", "User", "plan/method/return URLs", "Order + signed checkout", "201", "400/401/409", "Snapshot price/resources; server HMAC", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "POST", "/api/payments/sepay/ipn", "SePay", "API key + payload", "{success:true}", "200", "400/401/403/404/409", "Idempotent PAID + entitlement transaction", "payments.controller.ts -> PaymentsService", "SePay"),
    ("MF01", "GET", "/api/payments/:invoiceNumber", "User", "invoiceNumber", "PaymentOrder", "200", "401/403/404", "Owner only; reconcile pending", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "GET", "/api/payments/history", "User", "—", "PaymentOrder[]", "200", "401/403", "Owner only; normalize expiry", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "POST", "/api/payments/:invoiceNumber/status", "User", "FAILED/CANCELLED", "PaymentOrder", "200", "400/401/403/404/409", "PAID immutable by client", "payments.controller.ts -> PaymentsService", "SubscriptionView"),
    ("MF01", "INTERNAL", "SMTP email transport", "System", "from,to,subject,html", "Provider accepted", "N/A", "timeout/auth/TLS/5xx", "SMTP config qua secret; TLS; bounded retry; no secret/action-code logs; hiện source dùng Resend HTTPS nên đây là gap", "mail/mail.service.ts", "AuthEmailService"),
    ("MF02", "POST", "/api/documents", "User", "multipart file + metadata", "DocumentDto", "201", "400/413/415/422/429", "Quota/type/size/taxonomy; enqueue extraction", "documents.controller.ts -> DocumentsService", "UploadDocumentView"),
    ("MF02", "POST", "/api/documents/:id/extract", "User", "document UUID", "Job/status", "202", "403/404/409/429", "Access; no duplicate job; retry bound", "document-content.controller.ts", "Library retry"),
    ("MF02", "GET", "/api/documents/:id/content", "User", "document UUID", "Extracted content", "200", "403/404/409", "Accessible and COMPLETED", "document-content.controller.ts", "Document/chat UI"),
    ("MF02", "GET", "/api/documents/:id/extraction-status", "User", "document UUID", "Status/progress", "200", "403/404", "Safe status/error metadata", "document-content.controller.ts", "Library polling"),
    ("MF02", "GET", "/api/documents/:id/preview", "User", "document UUID", "Signed URL", "200", "403/404/422", "Access; short expiry; office fallback", "documents.controller.ts", "LibraryView"),
    ("MF02", "GET", "/api/documents/:id/download", "User", "document UUID", "Signed URL", "200", "403/404", "Access; count/log download", "documents.controller.ts", "LibraryView"),
    ("MF03", "POST", "/api/chat/ask-document", "User", "question/documentId/sessionId?", "Answer/sources", "200", "400/403/404/409/429", "Accessible completed doc; AI entitlement; grounded", "ai-chatbot.controller.ts", "AskDocumentView"),
    ("MF03", "POST", "/api/chat/ask-library", "User", "question/filters/sessionId?", "Answer/sources", "200", "400/403/409/429", "Owned+saved accessible scope", "ai-chatbot.controller.ts", "AskLibraryView"),
    ("MF03", "POST", "/api/chat/ask-document/stream", "User", "ask-document DTO", "SSE", "200", "Auth/error events", "Same auth/quota/persistence as non-stream", "ai-chatbot.controller.ts", "Chat UI"),
    ("MF03", "POST", "/api/chat/ask-library/stream", "User", "ask-library DTO", "SSE", "200", "Auth/error events", "Ordered status/sources/delta/done", "ai-chatbot.controller.ts", "Chat UI"),
    ("MF03", "GET", "/api/chat/sessions", "User", "query", "Sessions + meta", "200", "400/401", "Current user only", "ai-chatbot.controller.ts", "AiChatbotView"),
    ("MF03", "GET", "/api/chat/sessions/:id", "User", "session UUID", "Session", "200", "403/404", "Owner only", "ai-chatbot.controller.ts", "AiChatbotView"),
    ("MF03", "GET", "/api/chat/messages/:sessionId", "User", "pagination", "Messages + meta", "200", "403/404", "Owner; stable ordering", "ai-chatbot.controller.ts", "AiChatbotView"),
    ("MF04", "PUT", "/api/documents/:id/visibility", "Owner", "PRIVATE/PUBLIC", "Document", "200", "400/403/404", "PUBLIC -> PENDING", "documents.controller.ts", "LibraryView"),
    ("MF04", "GET", "/api/community/documents", "Optional auth", "query/page/limit", "Documents + meta", "200", "400", "PUBLIC+APPROVED+ACTIVE only", "community.controller.ts", "CommunityView"),
    ("MF04", "GET", "/api/community/documents/:id", "Optional auth", "UUID", "Document detail", "200", "404", "Same publication predicate", "community.controller.ts", "CommunityView"),
    ("MF04", "GET", "/api/community/documents/:id/preview", "User", "UUID", "Signed URL", "200", "401/403/404", "Published document only", "community.controller.ts", "CommunityView"),
    ("MF04", "POST", "/api/community/documents/:id/save", "User", "UUID", "Saved relation", "200/201", "403/404", "Idempotent; accessible document", "community.controller.ts", "CommunityView"),
    ("MF04", "DELETE", "/api/community/documents/:id/save", "User", "UUID", "No content", "204", "403/404", "Idempotent; own relation", "community.controller.ts", "Community/SavedView"),
    ("MF04", "GET", "/api/saved-documents", "User", "query/page/limit", "Documents + meta", "200", "400/401", "Own accessible saved docs", "saved-documents.controller.ts", "SavedView"),
    ("MF05", "GET", "/api/admin/documents", "Admin", "moderation filters/page/limit", "Documents + meta", "200", "400/401/403", "PUBLIC queue; prioritized", "admin-documents.controller.ts", "AdminDocumentsView"),
    ("MF05", "GET", "/api/admin/documents/:id/preview", "Admin", "UUID", "Signed URL", "200", "403/404", "PUBLIC; no state change", "admin-documents.controller.ts", "AdminDocumentsView"),
    ("MF05", "PUT", "/api/admin/documents/:id/approve", "Admin", "UUID", "Moderation result", "200", "403/404", "APPROVED+ACTIVE; audit+notify", "admin-documents.controller.ts", "AdminDocumentsView"),
    ("MF05", "PUT", "/api/admin/documents/:id/reject", "Admin", "reason", "Moderation result", "200", "400/403/404", "Reason required; REJECTED+HIDDEN; audit+notify", "admin-documents.controller.ts", "AdminDocumentsView"),
    ("MF05", "GET", "/api/admin/users", "Admin", "filters/page/limit", "Users + meta", "200", "400/401/403", "No credentials; DB role", "admin-users.controller.ts", "AdminUsersView"),
    ("MF05", "PATCH", "/api/admin/users/:id/status", "Admin", "status", "User", "200", "400/403/404/409", "Audit status transition", "admin-users.controller.ts", "AdminUsersView"),
    ("MF05", "GET", "/api/admin/logs/audit", "Admin", "filters/page/limit", "Audit + meta", "200", "400/401/403", "Append-only business audit", "audit-log.controller.ts", "Admin dashboard"),
    ("MF05", "GET/PATCH", "/api/notifications", "User", "limit/read operations", "Notifications", "200", "401/403/404", "Owner only; mark read idempotent", "notifications.controller.ts", "AppLayout"),
]


def base_style(ws, columns, widths):
    navy, blue = "17365D", "D9EAF7"
    thin = Side(style="thin", color="B7C9D6")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(columns)}{ws.max_row}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True, size=14)
    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=navy, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=columns):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:4"


def build_flow():
    wb = Workbook()
    ov = wb.active
    ov.title = "Tổng quan"
    ov.append(["DOCUMIND v0.4 — 5 MAIN FLOW & BUSINESS RULE"])
    ov.append(["Baseline", "Documind_5_main_flows_mapping.xlsx", "Version", "0.4", "Ngày", "2026-08-26", "Tổng rule", sum(len(f["steps"]) for f in FLOWS.values()), "Phân công", "Theo reconstruction guide"])
    ov.append(["Lưu ý MF01", "Email nghiệp vụ đã có qua Resend HTTPS; SMTP được ghi là gap cần bổ sung. Mỗi MF giữ nguyên owner và được bổ sung phạm vi Flutter tương ứng."])
    ov.merge_cells("B3:J3")
    ov.append(["Main Flow", "Tên luồng", "Thành viên", "Số bước/rule", "API chính", "Điểm chốt", "Mobile scope", "Mobile source", "SMTP", "Version"])
    for code, flow in FLOWS.items():
        member, mobile_scope, mobile_source = MEMBERS[code]
        ov.append([code, flow["name"], member, len(flow["steps"]), flow["steps"][0][3], flow["steps"][-1][4], mobile_scope, mobile_source, "Cần bổ sung" if code == "MF01" else "N/A", "v0.4"])
    base_style(ov, 10, [12, 40, 18, 14, 38, 52, 55, 50, 18, 12])
    for code, flow in FLOWS.items():
        member, mobile_scope, mobile_source = MEMBERS[code]
        ws = wb.create_sheet(code)
        ws.append([f"{code} — {flow['name'].upper()}"])
        ws.append(["Nguồn", "Documind_5_main_flows_mapping.xlsx", "Version", "0.4", "Thành viên", member, "Mobile source", mobile_source])
        ws.append(["Invariant", f"Giữ nguyên phạm vi {code}; {member} phụ trách cả Web/API mapping và phần Mobile App tương ứng. Business rule phải enforce ở backend/service/DB."])
        ws.merge_cells("B3:J3")
        ws.append(FLOW_HEADERS)
        for row in flow["steps"]:
            ws.append([*row, member, f"{mobile_scope}. Source: {mobile_source}"])
        base_style(ws, 10, [14, 30, 17, 45, 65, 45, 32, 20, 18, 65])
    wb.save(FLOW_OUT)


def build_api():
    wb = Workbook()
    ov = wb.active
    ov.title = "Tổng quan"
    ov.append(["DOCUMIND API CONTRACT v0.4 — MAPPED TO 5 MAIN FLOW"])
    ov.append(["Baseline", "Documind_5_main_flows_mapping.xlsx", "Base path", "/api", "Swagger", "/api/docs", "API rows", len(API_ROWS), "Version", "0.4", "Ngày", "2026-08-26", "Mobile", "Theo owner MF", "Scope", "MF01-MF05"])
    ov.append(["Envelope", "JSON: {success,data,timestamp}; pagination thêm meta; 204 không body; SSE/IPN/binary là ngoại lệ. SMTP là internal contract và hiện là implementation gap."])
    ov.merge_cells("B3:O3")
    ov.append(API_HEADERS)
    counts = {code: 0 for code in FLOWS}
    for row in API_ROWS:
        counts[row[0]] += 1
        member, mobile_scope, mobile_source = MEMBERS[row[0]]
        mobile_status = "Gap SMTP" if row[2] == "SMTP email transport" else "Mapped theo MF"
        ov.append([f"{row[0]}-API{counts[row[0]]:02d}", *row, member, f"{mobile_scope}; {mobile_source}", mobile_status])
    base_style(ov, 15, [15, 12, 12, 42, 18, 34, 30, 14, 34, 54, 42, 28, 18, 58, 18])
    for code, flow in FLOWS.items():
        member, mobile_scope, mobile_source = MEMBERS[code]
        ws = wb.create_sheet(code)
        ws.append([f"API CONTRACT {code} — {flow['name'].upper()}"])
        ws.append(["Main Flow", code, "Version", "0.4", "Thành viên", member, "Mobile source", mobile_source])
        ws.append(["Quy ước", "Firebase Bearer cho protected API; DB role/status; UUID; camelCase; ISO-8601; stable errors; requestId."])
        ws.merge_cells("B3:O3")
        ws.append(API_HEADERS)
        idx = 0
        for row in API_ROWS:
            if row[0] == code:
                idx += 1
                mobile_status = "Gap SMTP" if row[2] == "SMTP email transport" else "Mapped theo MF"
                ws.append([f"{code}-API{idx:02d}", *row, member, f"{mobile_scope}; {mobile_source}", mobile_status])
        base_style(ws, 15, [15, 12, 12, 42, 18, 34, 30, 14, 34, 54, 42, 28, 18, 58, 18])
    wb.save(API_OUT)


def verify():
    flow = load_workbook(FLOW_OUT, read_only=True)
    api = load_workbook(API_OUT, read_only=True)
    assert flow.sheetnames == ["Tổng quan", *FLOWS.keys()]
    assert api.sheetnames == ["Tổng quan", *FLOWS.keys()]
    assert sum(flow[c].max_row - 4 for c in FLOWS) == sum(len(x["steps"]) for x in FLOWS.values())
    assert sum(api[c].max_row - 4 for c in FLOWS) == len(API_ROWS)
    flow.close()
    api.close()
    print(FLOW_OUT)
    print(API_OUT)
    print(f"business_rules={sum(len(x['steps']) for x in FLOWS.values())} api_rows={len(API_ROWS)}")


if __name__ == "__main__":
    build_flow()
    build_api()
    verify()
